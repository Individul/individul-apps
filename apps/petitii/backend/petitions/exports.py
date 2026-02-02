from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io


def export_petitions_xlsx(queryset):
    """Export petitions to XLSX format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registru Petiții"

    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Nr. Înreg.",
        "Data Înreg.",
        "Tip Petiționar",
        "Nume Petiționar",
        "Nume Deținut",
        "Obiect",
        "Status",
        "Termen",
        "Atribuit",
        "Data Soluție",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row, petition in enumerate(queryset, 2):
        data = [
            petition.registration_number,
            petition.registration_date.strftime("%d.%m.%Y"),
            petition.get_petitioner_type_display(),
            petition.petitioner_name,
            petition.detainee_fullname or "-",
            petition.get_object_type_display(),
            petition.get_status_display(),
            petition.response_due_date.strftime("%d.%m.%Y"),
            petition.assigned_to.get_full_name() if petition.assigned_to else "-",
            petition.resolution_date.strftime("%d.%m.%Y") if petition.resolution_date else "-",
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Set column widths
    column_widths = [12, 12, 15, 25, 25, 20, 15, 12, 20, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"registru_petitii_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_petitions_pdf(queryset):
    """Export petitions to PDF format (official register style)."""
    buffer = io.BytesIO()

    # Create document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,  # Center
        spaceAfter=20
    )

    # Add title
    title = Paragraph("REGISTRU PETIȚII", title_style)
    elements.append(title)

    date_str = timezone.now().strftime("%d.%m.%Y")
    subtitle = Paragraph(f"Generat la: {date_str}", styles['Normal'])
    elements.append(subtitle)
    elements.append(Spacer(1, 20))

    # Table data
    table_data = [
        ["Nr.", "Nr. Înreg.", "Data", "Tip Pet.", "Petiționar", "Deținut", "Obiect", "Status", "Termen", "Soluție"]
    ]

    for idx, petition in enumerate(queryset, 1):
        row = [
            str(idx),
            petition.registration_number,
            petition.registration_date.strftime("%d.%m.%Y"),
            petition.get_petitioner_type_display()[:10],
            petition.petitioner_name[:20] + "..." if len(petition.petitioner_name) > 20 else petition.petitioner_name,
            (petition.detainee_fullname[:15] + "..." if petition.detainee_fullname and len(petition.detainee_fullname) > 15 else petition.detainee_fullname) or "-",
            petition.get_object_type_display()[:15],
            petition.get_status_display(),
            petition.response_due_date.strftime("%d.%m.%Y"),
            petition.resolution_date.strftime("%d.%m.%Y") if petition.resolution_date else "-",
        ]
        table_data.append(row)

    # Create table
    col_widths = [0.8*cm, 2.5*cm, 2.2*cm, 2*cm, 4*cm, 3.5*cm, 3*cm, 2.5*cm, 2.2*cm, 2.2*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.85, 0.85, 0.85)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Nr. column
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Date column
        ('ALIGN', (8, 1), (9, -1), 'CENTER'),  # Term and solution columns
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Build PDF
    doc.build(elements)

    # Create response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"registru_petitii_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
