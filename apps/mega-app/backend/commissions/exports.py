from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import io

from .models import Article, ProgramResult, BehaviorResult, Decision


MONTH_NAMES_RO = [
    '', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
    'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie'
]


def _get_title(year, month=None, quarter=None):
    if quarter:
        return f"RAPORT COMISIE PENITENCIARA - Trimestrul {quarter}/{year}"
    elif month:
        return f"RAPORT COMISIE PENITENCIARA - {MONTH_NAMES_RO[month]} {year}"
    else:
        return f"RAPORT COMISIE PENITENCIARA - Anul {year}"


def _get_aggregated_data(queryset):
    """Agregate date per articol din CommissionArticleResult queryset."""
    rows = []
    for art_choice in Article.choices:
        art_value, art_label = art_choice
        art_qs = queryset.filter(article=art_value)
        rows.append({
            'article': art_value,
            'article_display': art_label,
            'total': art_qs.count(),
            'realizat': art_qs.filter(program_result=ProgramResult.REALIZAT).count(),
            'nerealizat': art_qs.filter(program_result=ProgramResult.NEREALIZAT).count(),
            'nerealizat_independent': art_qs.filter(program_result=ProgramResult.NEREALIZAT_INDEPENDENT).count(),
            'pozitiv': art_qs.filter(behavior_result=BehaviorResult.POZITIV).count(),
            'negativ': art_qs.filter(behavior_result=BehaviorResult.NEGATIV).count(),
            'admis': art_qs.filter(decision=Decision.ADMIS).count(),
            'respins': art_qs.filter(decision=Decision.RESPINS).count(),
        })
    return rows


def export_commissions_xlsx(queryset, year=None, month=None, quarter=None):
    """Export raport comisie in XLSX."""
    wb = Workbook()
    ws = wb.active

    title = _get_title(year, month, quarter)
    ws.title = title[:31]

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    num_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bold_font = Font(bold=True, size=11)

    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        'Nr.', 'Articol', 'Total',
        'Realizat', 'Nerealizat', 'Nerealizat (ind.)',
        'Pozitiv', 'Negativ',
        'Admis', 'Respins',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    data_rows = _get_aggregated_data(queryset)

    totals = {
        'total': 0, 'realizat': 0, 'nerealizat': 0, 'nerealizat_independent': 0,
        'pozitiv': 0, 'negativ': 0, 'admis': 0, 'respins': 0,
    }

    for idx, row in enumerate(data_rows, 1):
        r = idx + 3
        ws.cell(row=r, column=1, value=idx).alignment = num_alignment
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=row['article_display']).alignment = cell_alignment
        ws.cell(row=r, column=2).border = thin_border

        values = [
            row['total'], row['realizat'], row['nerealizat'], row['nerealizat_independent'],
            row['pozitiv'], row['negativ'], row['admis'], row['respins'],
        ]
        for col_offset, val in enumerate(values):
            cell = ws.cell(row=r, column=col_offset + 3, value=val)
            cell.alignment = num_alignment
            cell.border = thin_border

        for key in totals:
            totals[key] += row[key]

    # Totals row
    r = len(data_rows) + 4
    ws.cell(row=r, column=1, value='').border = thin_border
    total_cell = ws.cell(row=r, column=2, value='TOTAL')
    total_cell.font = bold_font
    total_cell.alignment = cell_alignment
    total_cell.border = thin_border

    total_values = [
        totals['total'], totals['realizat'], totals['nerealizat'], totals['nerealizat_independent'],
        totals['pozitiv'], totals['negativ'], totals['admis'], totals['respins'],
    ]
    for col_offset, val in enumerate(total_values):
        cell = ws.cell(row=r, column=col_offset + 3, value=val)
        cell.font = bold_font
        cell.alignment = num_alignment
        cell.border = thin_border

    # Column widths
    widths = [5, 25, 8, 10, 12, 18, 10, 10, 10, 10]
    for col, width in enumerate(widths, 1):
        col_letter = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    suffix = f"{month or ''}{('T' + str(quarter)) if quarter else ''}"
    filename = f"comisie_{year}_{suffix or 'anual'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_commissions_pdf(queryset, year=None, month=None, quarter=None):
    """Export raport comisie in PDF."""
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=14, alignment=1, spaceAfter=20
    )

    title = _get_title(year, month, quarter)
    elements.append(Paragraph(title, title_style))

    date_str = timezone.now().strftime("%d.%m.%Y")
    elements.append(Paragraph(f"Generat la: {date_str}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Table data
    table_data = [[
        'Nr.', 'Articol', 'Total',
        'Realizat', 'Nerealizat', 'Nerealizat\n(ind.)',
        'Pozitiv', 'Negativ',
        'Admis', 'Respins',
    ]]

    data_rows = _get_aggregated_data(queryset)

    totals = {
        'total': 0, 'realizat': 0, 'nerealizat': 0, 'nerealizat_independent': 0,
        'pozitiv': 0, 'negativ': 0, 'admis': 0, 'respins': 0,
    }

    for idx, row in enumerate(data_rows, 1):
        table_data.append([
            str(idx),
            row['article_display'],
            str(row['total']),
            str(row['realizat']),
            str(row['nerealizat']),
            str(row['nerealizat_independent']),
            str(row['pozitiv']),
            str(row['negativ']),
            str(row['admis']),
            str(row['respins']),
        ])
        for key in totals:
            totals[key] += row[key]

    # Totals row
    table_data.append([
        '', 'TOTAL',
        str(totals['total']),
        str(totals['realizat']),
        str(totals['nerealizat']),
        str(totals['nerealizat_independent']),
        str(totals['pozitiv']),
        str(totals['negativ']),
        str(totals['admis']),
        str(totals['respins']),
    ])

    col_widths = [
        1 * cm, 3.5 * cm, 2 * cm,
        2.2 * cm, 2.5 * cm, 2.8 * cm,
        2.2 * cm, 2.2 * cm,
        2.2 * cm, 2.2 * cm,
    ]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.85, 0.85, 0.85)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        # Totals row bold
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
    ])

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    suffix = f"{month or ''}{('T' + str(quarter)) if quarter else ''}"
    filename = f"comisie_{year}_{suffix or 'anual'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
