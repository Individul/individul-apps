from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_issues_xlsx(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tracker SIA'

    headers = [
        'Titlu', 'Categorie', 'Prioritate', 'Status',
        'Modul afectat', 'Descriere', 'Creat de', 'Data creării',
    ]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row, issue in enumerate(queryset, 2):
        created_by_name = ''
        if issue.created_by:
            name = f"{issue.created_by.first_name} {issue.created_by.last_name}".strip()
            created_by_name = name or issue.created_by.username

        values = [
            issue.title,
            issue.get_category_display(),
            issue.get_priority_display(),
            issue.get_status_display(),
            issue.module_name,
            issue.description,
            created_by_name,
            issue.created_at.strftime('%d.%m.%Y %H:%M'),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tracker_sia.xlsx"'
    wb.save(response)
    return response
