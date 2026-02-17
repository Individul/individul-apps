from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import io

from .models import Penitentiary, ISOLATOR_VALUES


MONTH_NAMES_RO = [
    '', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
    'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie'
]


def _get_title(year, month=None, quarter=None):
    if quarter:
        return f"SITUATIA TRANSFERURILOR - Trimestrul {quarter}/{year}"
    elif month:
        return f"SITUATIA TRANSFERURILOR - {MONTH_NAMES_RO[month]} {year}"
    else:
        return f"SITUATIA TRANSFERURILOR - Anul {year}"


def _get_aggregated_data(queryset):
    """Agregate date per penitenciar din TransferEntry queryset."""
    rows = queryset.values('penitentiary').annotate(
        total_veniti=Sum('veniti'),
        total_veniti_reintorsi=Sum('veniti_reintorsi'),
        total_veniti_noi=Sum('veniti_noi'),
        total_plecati=Sum('plecati'),
        total_plecati_izolator=Sum('plecati_izolator'),
    ).order_by('penitentiary')

    result = []
    for row in rows:
        pen_val = row['penitentiary']
        result.append({
            'penitentiary': pen_val,
            'penitentiary_display': Penitentiary(pen_val).label,
            'veniti': row['total_veniti'] or 0,
            'veniti_reintorsi': row['total_veniti_reintorsi'] or 0,
            'veniti_noi': row['total_veniti_noi'] or 0,
            'plecati': row['total_plecati'] or 0,
            'plecati_izolator': row['total_plecati_izolator'] or 0,
        })
    return result


def export_transfers_xlsx(queryset, year=None, month=None, quarter=None):
    """Export transferuri in XLSX."""
    wb = Workbook()
    ws = wb.active

    title = _get_title(year, month, quarter)
    ws.title = title[:31]  # max 31 chars for sheet name

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    num_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bold_font = Font(bold=True, size=11)

    # Title row
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ['Nr.', 'Penitenciar', 'Veniti', 'Reintorsi', 'Noi', 'Plecati', 'La izolator', 'Observatii']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    data_rows = _get_aggregated_data(queryset)

    totals = {'veniti': 0, 'veniti_reintorsi': 0, 'veniti_noi': 0, 'plecati': 0, 'plecati_izolator': 0}

    for idx, row in enumerate(data_rows, 1):
        r = idx + 3
        is_iso = row['penitentiary'] in ISOLATOR_VALUES

        ws.cell(row=r, column=1, value=idx).alignment = num_alignment
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=row['penitentiary_display']).alignment = cell_alignment
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3, value=row['veniti']).alignment = num_alignment
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=row['veniti_reintorsi']).alignment = num_alignment
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=5, value=row['veniti_noi']).alignment = num_alignment
        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=6, value=row['plecati']).alignment = num_alignment
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=7, value=row['plecati_izolator'] if is_iso else '-').alignment = num_alignment
        ws.cell(row=r, column=7).border = thin_border
        ws.cell(row=r, column=8, value='').alignment = cell_alignment
        ws.cell(row=r, column=8).border = thin_border

        for key in totals:
            totals[key] += row[key]

    # Totals row
    r = len(data_rows) + 4
    ws.cell(row=r, column=1, value='').border = thin_border
    total_cell = ws.cell(row=r, column=2, value='TOTAL')
    total_cell.font = bold_font
    total_cell.alignment = cell_alignment
    total_cell.border = thin_border

    for col, key in enumerate(['veniti', 'veniti_reintorsi', 'veniti_noi', 'plecati', 'plecati_izolator'], 3):
        cell = ws.cell(row=r, column=col, value=totals[key])
        cell.font = bold_font
        cell.alignment = num_alignment
        cell.border = thin_border

    ws.cell(row=r, column=8, value='').border = thin_border

    # Column widths
    widths = [5, 30, 10, 12, 10, 10, 12, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    suffix = f"{month or ''}{('T' + str(quarter)) if quarter else ''}"
    filename = f"transferuri_{year}_{suffix or 'anual'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_transfers_pdf(queryset, year=None, month=None, quarter=None):
    """Export transferuri in PDF."""
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=14, alignment=1, spaceAfter=20
    )

    title = _get_title(year, month, quarter)
    elements.append(Paragraph(title, title_style))

    date_str = timezone.now().strftime("%d.%m.%Y")
    elements.append(Paragraph(f"Generat la: {date_str}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Table data
    table_data = [['Nr.', 'Penitenciar', 'Veniti', 'Reintorsi', 'Noi', 'Plecati', 'La izolator']]

    data_rows = _get_aggregated_data(queryset)

    totals = {'veniti': 0, 'veniti_reintorsi': 0, 'veniti_noi': 0, 'plecati': 0, 'plecati_izolator': 0}

    for idx, row in enumerate(data_rows, 1):
        is_iso = row['penitentiary'] in ISOLATOR_VALUES
        table_data.append([
            str(idx),
            row['penitentiary_display'],
            str(row['veniti']),
            str(row['veniti_reintorsi']),
            str(row['veniti_noi']),
            str(row['plecati']),
            str(row['plecati_izolator']) if is_iso else '-',
        ])
        for key in totals:
            totals[key] += row[key]

    # Totals row
    table_data.append([
        '', 'TOTAL',
        str(totals['veniti']),
        str(totals['veniti_reintorsi']),
        str(totals['veniti_noi']),
        str(totals['plecati']),
        str(totals['plecati_izolator']),
    ])

    col_widths = [1 * cm, 6 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.85, 0.85, 0.85)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        # Totals row bold
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
    ])

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    suffix = f"{month or ''}{('T' + str(quarter)) if quarter else ''}"
    filename = f"transferuri_{year}_{suffix or 'anual'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
